"""
Sistema de Análise de Requerimentos - Gestão CSC (VERSÃO 3.2 - SEPARAÇÃO POR DATA)
===============================================================================

Versão 3.2 - NOVIDADE CRÍTICA:
- ✅ NÃO SOMA dados de datas diferentes
- ✅ Cada pasta/data gera análise SEPARADA
- ✅ Colaboradores são identificados por DATA + NOME
- ✅ Exemplo: "Andrey (05.02.2026)" e "Andrey (06.02.2026)" são distintos
- ✅ Relatório mostra evolução por data de forma clara

Versão 3.1 - FUNCIONALIDADES:
- ✅ Imagens PNG incorporadas diretamente no HTML (Base64)
- ✅ Relatório HTML autocontido - perfeito para email

Versão 3.0 - MUDANÇAS PRINCIPAIS:
- ✅ Usa data da PASTA (formato DD.MM.YYYY) como base temporal
- ✅ Estrutura de pastas: planilhas_para_analise/06.02.2026/arquivo.xlsm

Autor: Jonathan Barbosa 
Data: 2026-02-08
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
from datetime import datetime
import warnings
import glob
from typing import List, Dict, Tuple, Optional
import os
import shutil
import zipfile
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import base64
from io import BytesIO

warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURAÇÕES
# ============================================================================

class Config:
    """Classe de configuração centralizada"""
    
    PASTA_PLANILHAS = "./planilhas_para_analise"
    
    COL_SITUACAO = "SITUAÇÃO"
    COL_COLABORADOR = "COLABORADORES"
    COL_DATA_INICIO = "DT_INICIO_ETAPA"
    COL_DATA_PRAZO = "DT_PRAZO_FINAL"
    COL_PROTOCOLO = "PROTOCOLO"
    COL_TIPO_REQ = "TIPO_REQUERIMENTO"
    COL_CAMPUS = "NOM_CAMPUS"
    COL_CURSO = "NOM_CURSO"
    
    INDICE_COLUNA_SITUACAO = 8
    
    VALORES_RESPONDIDO = [
        "Deferido",
        "Indeferido", 
        "Redirecionado"
    ]
    
    CORES_FORMATACAO_VALIDAS = [
        'FF92D050',
        'FFFF0000',
        'FFFFFF00',
        'FF00B0F0',
    ]
    
    NOME_ABA = "BASE"
    PASTA_SAIDA = "./resultado_analise"
    
    ESTILO_GRAFICO = 'seaborn-v0_8-darkgrid'
    CORES_PADRAO = ['#00B3B0', '#061A2B', '#3EE7DA', '#0A2A44', '#16697A', 
                    '#489FB5', '#82C0CC', '#114B5F', '#2E8B9E', '#1A4A5C']


# ============================================================================
# CLASSE AUXILIAR PARA LEITURA DE FORMATAÇÃO
# ============================================================================

class LeitorFormatacaoCondicional:
    """Classe para ler formatação condicional de células Excel"""

    @staticmethod
    def _cor_para_argb_upper(color_obj) -> Optional[str]:
        if color_obj is None:
            return None
        rgb = getattr(color_obj, 'rgb', None)
        if rgb:
            rgb_str = str(rgb).upper()
            if len(rgb_str) == 6:
                return f"FF{rgb_str}"
            return rgb_str
        return None
    
    @staticmethod
    def ler_linhas_com_formatacao(arquivo: str, aba: str, config: Config) -> List[int]:
        """Lê linhas que possuem formatação condicional nas colunas A até I"""
        try:
            wb = load_workbook(arquivo, data_only=False)
            ws = wb[aba]
            
            linhas_formatadas = []
            cores_validas = {str(c).upper() for c in getattr(config, 'CORES_FORMATACAO_VALIDAS', []) if c}
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=9), start=1):
                tem_formatacao = False
                
                for cell in row:
                    if not cell.fill:
                        continue

                    pattern_type = getattr(cell.fill, 'patternType', None)

                    cor = LeitorFormatacaoCondicional._cor_para_argb_upper(getattr(cell.fill, 'fgColor', None))
                    if not cor:
                        cor = LeitorFormatacaoCondicional._cor_para_argb_upper(getattr(cell.fill, 'start_color', None))

                    if not cor:
                        if pattern_type and str(pattern_type).lower() != 'none':
                            tem_formatacao = True
                            break
                        continue

                    if cor == '00000000':
                        continue

                    if cores_validas:
                        if cor in cores_validas:
                            tem_formatacao = True
                            break
                    else:
                        tem_formatacao = True
                        break
                
                if tem_formatacao:
                    linhas_formatadas.append(row_idx - 1)
            
            wb.close()
            return linhas_formatadas
            
        except Exception as e:
            print(f"    ⚠ Erro ao ler formatação condicional: {str(e)}")
            return []
    
    @staticmethod
    def ler_valores_coluna_i(arquivo: str, aba: str) -> pd.Series:
        """Lê especificamente a coluna I de forma robusta"""
        try:
            wb = load_workbook(arquivo, data_only=True)
            ws = wb[aba]
            
            valores = []
            
            for row in ws.iter_rows(min_row=2, min_col=9, max_col=9, values_only=True):
                valores.append(row[0] if row[0] is not None else '')
            
            wb.close()
            return pd.Series(valores)
            
        except Exception as e:
            print(f"    ⚠ Erro ao ler coluna I: {str(e)}")
            return pd.Series()


# ============================================================================
# CLASSE PRINCIPAL DE ANÁLISE (VERSÃO 3.2 - SEPARAÇÃO POR DATA)
# ============================================================================

class AnalisadorRequerimentos:
    """Classe principal para análise de requerimentos - Versão 3.2 com separação por data"""
    
    def __init__(self, config: Config):
        self.config = config
        self.dados_consolidados = None
        self.resultados = {}
        self.log_detalhado = []
        self._valores_respondido_norm = {self._normalizar_texto(v) for v in self.config.VALORES_RESPONDIDO}

        self.imagens_base64 = {}

        self.pasta_saida_base = self.config.PASTA_SAIDA
        self.pasta_saida_execucao = os.path.join(
            self.pasta_saida_base,
            datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        )
        self.pasta_saida_ultimo = os.path.join(self.pasta_saida_base, 'ULTIMO')

        Path(self.pasta_saida_execucao).mkdir(parents=True, exist_ok=True)
        Path(self.pasta_saida_ultimo).mkdir(parents=True, exist_ok=True)
        
        print("=" * 80)
        print("SISTEMA DE ANÁLISE DE REQUERIMENTOS - VERSÃO 3.2 (SEPARAÇÃO POR DATA)")
        print("=" * 80)
        print(f"Pasta de entrada: {self.config.PASTA_PLANILHAS}")
        print(f"Pasta de saída (execução): {self.pasta_saida_execucao}")
        print(f"Pasta de saída (último): {self.pasta_saida_ultimo}")
        print("=" * 80)
        print("🔑 NOVIDADE v3.2: Colaboradores são separados por DATA")
        print("   Exemplo: Andrey (05.02) ≠ Andrey (06.02)")
        print("=" * 80)

    def _fig_to_base64(self, fig) -> str:
        """Converte uma figura matplotlib em string base64"""
        buffer = BytesIO()
        fig.savefig(buffer, format='png', dpi=150, bbox_inches='tight')
        buffer.seek(0)
        image_base64 = base64.b64encode(buffer.read()).decode('utf-8')
        buffer.close()
        plt.close(fig)
        return f"data:image/png;base64,{image_base64}"

    def _caminhos_saida(self, nome_arquivo: str) -> Tuple[str, str]:
        return (
            os.path.join(self.pasta_saida_execucao, nome_arquivo),
            os.path.join(self.pasta_saida_ultimo, nome_arquivo)
        )

    def _atualizar_ultimo(self, caminho_exec: str, caminho_ultimo: str) -> None:
        try:
            shutil.copy2(caminho_exec, caminho_ultimo)
        except Exception as e:
            print(f"  ⚠ Não foi possível atualizar ULTIMO: {type(e).__name__}: {str(e)}")
    
    def _normalizar_texto(self, texto: str) -> str:
        """Normaliza texto para comparação"""
        if pd.isna(texto):
            return ''
        return str(texto).strip().lower()

    def _normalizar_nome_colaborador(self, nome: str) -> str:
        if pd.isna(nome):
            return 'Desconhecido'

        s = str(nome).strip()
        if not s or s.lower() in {'nan', 'none'}:
            return 'Desconhecido'

        s = re.sub(r"\s*\([^)]*\)", "", s)
        s = re.sub(r"[^A-Za-zÀ-ÿ\s]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()

        if not s:
            return 'Desconhecido'

        primeiro = s.split(' ')[0]
        return primeiro.capitalize()
    
    def _eh_valor_respondido(self, valor: str) -> bool:
        """Verifica se um valor indica requerimento respondido"""
        valor_norm = self._normalizar_texto(valor)
        return valor_norm in self._valores_respondido_norm
    
    def _extrair_data_da_pasta(self, caminho_arquivo: str) -> Optional[datetime]:
        """Extrai a data do nome da pasta no formato DD.MM.YYYY"""
        try:
            pasta_pai = Path(caminho_arquivo).parent.name
            padrao = r'(\d{2})\.(\d{2})\.(\d{4})'
            match = re.search(padrao, pasta_pai)
            
            if match:
                dia, mes, ano = match.groups()
                data = datetime(int(ano), int(mes), int(dia))
                return data
            
            return None
            
        except Exception as e:
            print(f"    ⚠ Erro ao extrair data da pasta: {str(e)}")
            return None
    
    def carregar_planilhas(self) -> pd.DataFrame:
        """Carrega todas as planilhas Excel da pasta especificada"""
        print("\n🔍 CARREGANDO PLANILHAS (Modo v3.2 - Separação por Data)...")
        
        padroes = [
            os.path.join(self.config.PASTA_PLANILHAS, "**", "*.xlsx"),
            os.path.join(self.config.PASTA_PLANILHAS, "**", "*.xlsm")
        ]
        
        arquivos = []
        for padrao in padroes:
            arquivos.extend(glob.glob(padrao, recursive=True))
        
        if not arquivos:
            raise FileNotFoundError(
                f"❌ Nenhuma planilha encontrada em: {self.config.PASTA_PLANILHAS}"
            )
        
        print(f"✓ Encontrados {len(arquivos)} arquivo(s)")
        
        lista_dfs = []
        leitor_formatacao = LeitorFormatacaoCondicional()
        
        for arquivo in arquivos:
            try:
                nome_arquivo = Path(arquivo).name
                caminho_relativo = os.path.relpath(arquivo, self.config.PASTA_PLANILHAS)
                print(f"\n📄 Processando: {caminho_relativo}")
                
                data_pasta = self._extrair_data_da_pasta(arquivo)
                if data_pasta:
                    print(f"    📅 Data da pasta detectada: {data_pasta.strftime('%d.%m.%Y')}")
                else:
                    print(f"    ⚠ Não foi possível extrair data da pasta")
                
                try:
                    xls = pd.ExcelFile(arquivo, engine='openpyxl')
                except zipfile.BadZipFile:
                    print(f"    ❌ Arquivo inválido/corrompido: {nome_arquivo}")
                    continue
                except ImportError as e:
                    raise ImportError(
                        "❌ Dependência 'openpyxl' não encontrada. Instale com: pip install openpyxl"
                    ) from e
                
                abas_disponiveis = list(xls.sheet_names)
                if not abas_disponiveis:
                    raise ValueError("❌ Nenhuma aba encontrada no arquivo")
                
                aba_desejada = self.config.NOME_ABA
                
                def _normalizar_nome_aba(s: str) -> str:
                    return "".join(str(s).strip().lower().split())
                
                if aba_desejada not in abas_disponiveis:
                    alvo_norm = _normalizar_nome_aba(aba_desejada)
                    mapa_norm = {_normalizar_nome_aba(a): a for a in abas_disponiveis}
                    if alvo_norm in mapa_norm:
                        aba_desejada = mapa_norm[alvo_norm]
                    else:
                        print(f"    ⚠ Aba '{self.config.NOME_ABA}' não encontrada.")
                        print(f"      Usando primeira aba: '{abas_disponiveis[0]}'")
                        aba_desejada = abas_disponiveis[0]
                
                df = pd.read_excel(arquivo, sheet_name=aba_desejada, engine='openpyxl')
                print(f"    ✓ {len(df)} linhas carregadas")
                
                print(f"    🎨 Tentando ler formatação condicional...")
                linhas_formatadas = leitor_formatacao.ler_linhas_com_formatacao(
                    arquivo, aba_desejada, self.config
                )
                
                if linhas_formatadas:
                    print(f"    ✓ {len(linhas_formatadas)} linhas com formatação detectadas")
                    df['TEM_FORMATACAO'] = False
                    linhas_formatadas_validas = [i for i in linhas_formatadas if isinstance(i, (int, np.integer)) and 0 <= int(i) < len(df)]
                    if linhas_formatadas_validas:
                        df.loc[linhas_formatadas_validas, 'TEM_FORMATACAO'] = True
                else:
                    print(f"    ⚠ Nenhuma formatação detectada ou erro na leitura")
                    df['TEM_FORMATACAO'] = False
                
                print(f"    📊 Lendo coluna I (Situação)...")
                
                col_situacao = None

                def _score_coluna_situacao(serie: pd.Series) -> int:
                    try:
                        norm = serie.apply(self._normalizar_texto)
                        return int(norm.isin(self._valores_respondido_norm).sum())
                    except Exception:
                        return 0

                candidatos: List[Tuple[str, pd.Series]] = []

                colunas_possiveis = ['SUTUAÇÃO', 'SITUAÇÃO', 'SITUACAO', 'STATUS']
                for col in colunas_possiveis:
                    if col in df.columns:
                        candidatos.append((col, df[col]))

                if len(df.columns) > self.config.INDICE_COLUNA_SITUACAO:
                    col_i_df = df.columns[self.config.INDICE_COLUNA_SITUACAO]
                    candidatos.append((col_i_df, df[col_i_df]))

                valores_col_i = None
                try:
                    valores_col_i = leitor_formatacao.ler_valores_coluna_i(arquivo, aba_desejada)
                except Exception:
                    valores_col_i = None

                if valores_col_i is not None and (not valores_col_i.empty) and len(valores_col_i) == len(df):
                    df['SITUACAO_COLUNA_I'] = valores_col_i
                    candidatos.append(('SITUACAO_COLUNA_I', df['SITUACAO_COLUNA_I']))

                if not candidatos:
                    print(f"    ❌ ERRO: Não foi possível identificar coluna de situação!")
                    continue

                scores = [(nome, _score_coluna_situacao(serie)) for nome, serie in candidatos]
                melhor_col, melhor_score = max(scores, key=lambda x: x[1])
                col_situacao = melhor_col

                scores_str = ", ".join([f"{n}={s}" for n, s in scores])
                print(f"    ✓ Coluna de situação escolhida: '{col_situacao}' (matches={melhor_score}; {scores_str})")
                
                if col_situacao is None:
                    print(f"    ❌ ERRO: Não foi possível identificar coluna de situação!")
                    continue
                
                df['COLUNA_SITUACAO_USADA'] = col_situacao
                df['SITUACAO_ORIGINAL'] = df[col_situacao]
                df['SITUACAO_NORMALIZADA'] = df[col_situacao].apply(self._normalizar_texto)
                df['EH_RESPONDIDO'] = df['SITUACAO_NORMALIZADA'].apply(self._eh_valor_respondido)
                
                qtd_respondidos = df['EH_RESPONDIDO'].sum()
                print(f"    ✓ {qtd_respondidos} requerimentos RESPONDIDOS identificados")
                
                valores_unicos = df[col_situacao].value_counts()
                print(f"    📋 Valores únicos na coluna situação:")
                for valor, count in valores_unicos.head(10).items():
                    eh_resp = "✓" if self._eh_valor_respondido(valor) else "✗"
                    print(f"       {eh_resp} '{valor}': {count}")
                
                df['ARQUIVO_ORIGEM'] = nome_arquivo
                
                if data_pasta:
                    df['DATA_PASTA'] = data_pasta
                    # NOVIDADE v3.2: Criar string de data formatada para identificação
                    df['DATA_PASTA_STR'] = data_pasta.strftime('%d.%m.%Y')
                else:
                    df['DATA_PASTA'] = pd.NaT
                    df['DATA_PASTA_STR'] = 'Sem Data'
                
                nome_stem = Path(nome_arquivo).stem
                nome_lower = nome_stem.lower()
                nome_collab = 'Desconhecido'
                if nome_lower.startswith('gestao_requerimentos_'):
                    resto = nome_stem[len('gestao_requerimentos_'):]
                    tokens = [t.strip() for t in resto.split('_') if t.strip()]
                    for t in tokens:
                        if any(ch.isalpha() for ch in t):
                            nome_collab = t
                            break
                elif 'gestao' in nome_lower or 'requerimento' in nome_lower:
                    tokens = [t.strip() for t in nome_stem.split('_') if t.strip()]
                    for t in tokens[::-1]:
                        if any(ch.isalpha() for ch in t):
                            nome_collab = t
                            break

                df['COLABORADOR_ARQUIVO'] = str(nome_collab).strip().capitalize() if nome_collab else 'Desconhecido'
                df['COLABORADOR_ARQUIVO'] = df['COLABORADOR_ARQUIVO'].apply(self._normalizar_nome_colaborador)
                
                lista_dfs.append(df)
                
                self.log_detalhado.append({
                    'arquivo': nome_arquivo,
                    'caminho_relativo': caminho_relativo,
                    'data_pasta': data_pasta.strftime('%d.%m.%Y') if data_pasta else 'N/A',
                    'linhas_total': len(df),
                    'linhas_formatadas': len(linhas_formatadas) if linhas_formatadas else 0,
                    'respondidos': qtd_respondidos,
                    'coluna_situacao': col_situacao
                })
                
            except Exception as e:
                print(f"    ❌ Erro ao processar {nome_arquivo}:")
                print(f"       {type(e).__name__}: {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        if not lista_dfs:
            raise ValueError("❌ Nenhuma planilha foi carregada com sucesso!")
        
        df_consolidado = pd.concat(lista_dfs, ignore_index=True)
        
        print(f"\n✅ CONSOLIDAÇÃO CONCLUÍDA")
        print(f"   📊 Total de registros: {len(df_consolidado)}")
        print(f"   ✓ Total de RESPONDIDOS: {df_consolidado['EH_RESPONDIDO'].sum()}")
        
        registros_com_data = df_consolidado['DATA_PASTA'].notna().sum()
        print(f"   📅 Registros com DATA_PASTA: {registros_com_data} de {len(df_consolidado)}")
        
        # NOVIDADE v3.2: Mostrar datas únicas encontradas
        if 'DATA_PASTA_STR' in df_consolidado.columns:
            datas_unicas = df_consolidado['DATA_PASTA_STR'].unique()
            print(f"   🗓️  Datas identificadas: {', '.join(sorted(datas_unicas))}")
        
        self.dados_consolidados = df_consolidado
        return df_consolidado
    
    def processar_dados(self) -> pd.DataFrame:
        """Processa e limpa os dados consolidados"""
        print("\n⚙️  PROCESSANDO DADOS (v3.2 - Separação por Data)...")
        
        df = self.dados_consolidados.copy()
        
        colunas_data = [self.config.COL_DATA_INICIO, self.config.COL_DATA_PRAZO]
        for col in colunas_data:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
                print(f"  ✓ Coluna '{col}' convertida para data (mantida para compatibilidade)")
        
        if 'DATA_PASTA' in df.columns:
            df['DATA_PROCESSAMENTO'] = df['DATA_PASTA']
            print(f"  ✅ Usando DATA_PASTA como base para análise temporal")
        else:
            print(f"  ⚠ DATA_PASTA não encontrada! Fallback para DT_INICIO_ETAPA")
            if self.config.COL_DATA_INICIO in df.columns:
                df['DATA_PROCESSAMENTO'] = df[self.config.COL_DATA_INICIO]
        
        df_respondidos = df[df['EH_RESPONDIDO'] == True].copy()

        col_planilha = self.config.COL_COLABORADOR
        if col_planilha in df_respondidos.columns:
            col_valida = df_respondidos[col_planilha].astype(str).str.strip()
            col_valida = col_valida.replace({'': np.nan, 'nan': np.nan, 'None': np.nan, 'NAN': np.nan})
            df_respondidos['COLABORADOR_FINAL'] = col_valida
            if 'COLABORADOR_ARQUIVO' in df_respondidos.columns:
                df_respondidos['COLABORADOR_FINAL'] = df_respondidos['COLABORADOR_FINAL'].fillna(df_respondidos['COLABORADOR_ARQUIVO'])
        elif 'COLABORADOR_ARQUIVO' in df_respondidos.columns:
            df_respondidos['COLABORADOR_FINAL'] = df_respondidos['COLABORADOR_ARQUIVO']

        if 'COLABORADOR_FINAL' in df_respondidos.columns:
            df_respondidos['COLABORADOR_FINAL'] = df_respondidos['COLABORADOR_FINAL'].apply(self._normalizar_nome_colaborador)
        
        # ============================================================
        # NOVIDADE v3.2: CRIAR COLABORADOR COM DATA
        # ============================================================
        if 'COLABORADOR_FINAL' in df_respondidos.columns and 'DATA_PASTA_STR' in df_respondidos.columns:
            df_respondidos['COLABORADOR_COM_DATA'] = df_respondidos.apply(
                lambda row: f"{row['COLABORADOR_FINAL']} ({row['DATA_PASTA_STR']})", 
                axis=1
            )
            print(f"  🔑 Criada coluna COLABORADOR_COM_DATA (separação por data)")
        
        print(f"  ✅ Filtrados {len(df_respondidos)} requerimentos RESPONDIDOS")
        print(f"     (de {len(df)} registros totais)")
        
        if len(df_respondidos) == 0:
            print(f"\n  ⚠️  AVISO: Nenhum requerimento respondido encontrado!")
        
        self.dados_processados = df_respondidos
        return df_respondidos
    
    def calcular_kpis(self) -> Dict:
        """Calcula os principais KPIs com separação por data"""
        print("\n📊 CALCULANDO KPIs (v3.2 - Separado por Data)...")
        
        df = self.dados_processados
        
        kpis = {}
        
        kpis['total_respondidos'] = len(df)
        print(f"  ✅ Total EXATO de respondidos: {kpis['total_respondidos']}")
        
        # ============================================================
        # NOVIDADE v3.2: USAR COLABORADOR_COM_DATA
        # ============================================================
        if 'COLABORADOR_COM_DATA' in df.columns:
            kpis['por_colaborador_com_data'] = df['COLABORADOR_COM_DATA'].value_counts()
            print(f"  ✅ Análise por colaborador COM DATA concluída:")
            for collab, qtd in kpis['por_colaborador_com_data'].items():
                print(f"     • {collab}: {qtd} requerimentos")
        
        # Manter análise sem data para compatibilidade
        col_collab = 'COLABORADOR_FINAL' if 'COLABORADOR_FINAL' in df.columns else None
        if col_collab:
            kpis['por_colaborador'] = df[col_collab].value_counts()
        
        if 'por_colaborador_com_data' in kpis and len(kpis['por_colaborador_com_data']) > 0:
            kpis['top_colaborador'] = kpis['por_colaborador_com_data'].index[0]
            kpis['top_colaborador_qtd'] = kpis['por_colaborador_com_data'].iloc[0]
            print(f"  🏆 Top colaborador+data: {kpis['top_colaborador']} ({kpis['top_colaborador_qtd']} req)")
        
        if 'DATA_PROCESSAMENTO' in df.columns:
            df_com_data = df.dropna(subset=['DATA_PROCESSAMENTO'])
            
            if not df_com_data.empty:
                kpis['evolucao_diaria'] = df_com_data.groupby(
                    df_com_data['DATA_PROCESSAMENTO'].dt.date
                ).size()
                
                kpis['evolucao_semanal'] = df_com_data.groupby(
                    df_com_data['DATA_PROCESSAMENTO'].dt.to_period('W')
                ).size()
                
                dias_trabalhados = len(kpis['evolucao_diaria'])
                kpis['media_dia'] = kpis['total_respondidos'] / dias_trabalhados if dias_trabalhados > 0 else 0
                
                print(f"  ✅ Média diária: {kpis['media_dia']:.1f} requerimentos/dia")
                print(f"  📅 Período: {df_com_data['DATA_PROCESSAMENTO'].min().date()} a {df_com_data['DATA_PROCESSAMENTO'].max().date()}")
        
        if 'SITUACAO_ORIGINAL' in df.columns:
            df_situacao = df[df['EH_RESPONDIDO'] == True]
            kpis['por_situacao'] = df_situacao['SITUACAO_ORIGINAL'].value_counts()
            print(f"  ✅ Distribuição por situação:")
            for situacao, qtd in kpis['por_situacao'].items():
                print(f"     • {situacao}: {qtd}")
        
        if self.config.COL_TIPO_REQ in df.columns:
            kpis['por_tipo'] = df[self.config.COL_TIPO_REQ].value_counts().head(10)
            print(f"  ✅ Top 10 tipos de requerimento identificados")
        
        self.resultados = kpis
        return kpis
    
    def gerar_graficos(self):
        """Gera todos os gráficos de análise"""
        print("\n📈 GERANDO GRÁFICOS (v3.2 - com separação por data)...")
        
        plt.style.use(self.config.ESTILO_GRAFICO)
        
        if 'por_colaborador_com_data' in self.resultados:
            self._grafico_barras_colaborador_com_data()
        
        if 'evolucao_diaria' in self.resultados:
            self._grafico_evolucao_temporal_barras()
        
        if 'por_situacao' in self.resultados:
            self._grafico_pizza_situacao()
        
        if 'por_tipo' in self.resultados:
            self._grafico_top_tipos()
        
        print("  ✅ Todos os gráficos gerados (PNG + Base64)!")
    
    def _grafico_barras_colaborador_com_data(self):
        """Gráfico de barras: requerimentos por colaborador COM DATA"""
        fig, ax = plt.subplots(figsize=(14, max(8, len(self.resultados['por_colaborador_com_data']) * 0.4)))
        
        dados = self.resultados['por_colaborador_com_data'].sort_values(ascending=True)
        cores = [self.config.CORES_PADRAO[i % len(self.config.CORES_PADRAO)] for i in range(len(dados))]
        
        dados.plot(kind='barh', ax=ax, color=cores, edgecolor='black')
        
        ax.set_title('Requerimentos Respondidos por Colaborador (Separado por Data)', 
                     fontsize=16, fontweight='bold', pad=20)
        ax.set_xlabel('Quantidade de Requerimentos', fontsize=12, fontweight='bold')
        ax.set_ylabel('Colaborador (Data)', fontsize=12, fontweight='bold')
        ax.grid(axis='x', alpha=0.3)
        
        for i, v in enumerate(dados.values):
            ax.text(v + 0.5, i, str(v), va='center', fontweight='bold', fontsize=10)
        
        plt.tight_layout()
        
        caminho_exec, caminho_ultimo = self._caminhos_saida('grafico_colaboradores_por_data.png')
        plt.savefig(caminho_exec, dpi=300, bbox_inches='tight')
        self._atualizar_ultimo(caminho_exec, caminho_ultimo)
        
        self.imagens_base64['grafico_colaboradores'] = self._fig_to_base64(fig)
        
        print(f"  ✓ Gráfico de colaboradores POR DATA (PNG + Base64)")
    
    def _grafico_evolucao_temporal_barras(self):
        """Gráfico de evolução temporal"""
        evolucao_diaria = self.resultados['evolucao_diaria'].sort_index()

        datas = pd.to_datetime(pd.Index(evolucao_diaria.index), errors='coerce')
        evolucao_series = pd.Series(evolucao_diaria.values, index=datas).dropna()
        
        if evolucao_series.empty:
            print("  ⚠ Sem dados temporais para gráfico")
            return
        
        dias_totais = (evolucao_series.index.max() - evolucao_series.index.min()).days + 1
        dias_com_dados = len(evolucao_series)
        total_geral = float(evolucao_series.sum())
        media_diaria = total_geral / dias_com_dados if dias_com_dados > 0 else 0
        
        if dias_com_dados <= 10:
            labels = []
            dias_semana = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
            for d in evolucao_series.index:
                dia_nome = dias_semana[d.weekday()]
                labels.append(f"{dia_nome}\n{d.strftime('%d.%m')}")
            valores = list(evolucao_series.values)
            titulo = f'Produção Diária ({dias_com_dados} dias)'
            subtitulo = f'Média: {media_diaria:.1f} req/dia | Separado por Data'
            
        elif dias_com_dados <= 60:
            evolucao_semanal = evolucao_series.groupby(
                evolucao_series.index.to_period('W')
            ).sum().sort_index()
            
            labels = []
            for i, p in enumerate(evolucao_semanal.index, 1):
                inicio = p.start_time.strftime('%d.%m')
                fim = p.end_time.strftime('%d.%m')
                labels.append(f"Semana {i}\n{inicio}–{fim}")
            
            valores = list(evolucao_semanal.values)
            titulo = f'Produção Semanal ({len(valores)} semanas)'
            subtitulo = f'Total: {total_geral:.0f} req | Média: {media_diaria:.1f} req/dia'
            
        else:
            evolucao_mensal = evolucao_series.groupby(
                evolucao_series.index.to_period('M')
            ).sum().sort_index()
            
            meses = {
                1: 'Jan', 2: 'Fev', 3: 'Mar', 4: 'Abr', 5: 'Mai', 6: 'Jun',
                7: 'Jul', 8: 'Ago', 9: 'Set', 10: 'Out', 11: 'Nov', 12: 'Dez'
            }
            
            labels = []
            for p in evolucao_mensal.index:
                mes_nome = meses[p.month]
                labels.append(f"{mes_nome}/{p.year}")
            
            valores = list(evolucao_mensal.values)
            titulo = f'Produção Mensal ({len(valores)} meses)'
            subtitulo = f'Total: {total_geral:.0f} req | Média: {media_diaria:.1f} req/dia'
        
        fig, ax = plt.subplots(figsize=(14, max(7, len(valores) * 0.6)))
        
        y_pos = np.arange(len(valores))
        cores = ['#00B3B0'] * len(valores)
        
        bars = ax.barh(y_pos, valores, color=cores, edgecolor='#061A2B', linewidth=1.2)
        
        for i, (bar, valor) in enumerate(zip(bars, valores)):
            width = bar.get_width()
            if width > max(valores) * 0.15:
                pos_x = width * 0.5
                ha = 'center'
                color = 'white'
            else:
                pos_x = width + max(valores) * 0.02
                ha = 'left'
                color = '#00B3B0'
            
            ax.text(pos_x, bar.get_y() + bar.get_height()/2, 
                   f'{int(valor)}',
                   ha=ha, va='center', fontweight='bold', fontsize=11, color=color)
        
        if dias_totais > 7:
            media_periodo = np.mean(valores)
            ax.axvline(media_periodo, color='#3EE7DA', linestyle='--', 
                      linewidth=2, alpha=0.7, label=f'Média: {media_periodo:.1f}')
            ax.legend(loc='lower right', fontsize=10, framealpha=0.9)
        
        ax.set_yticks(y_pos)
        ax.set_yticklabels(labels, fontsize=10)
        ax.set_xlabel('Quantidade de Requerimentos', fontsize=12, fontweight='bold', color='#00B3B0')
        ax.set_title(titulo, fontsize=16, fontweight='bold', color='#00B3B0', pad=10)
        
        ax.text(0.5, 1.02, subtitulo, 
                transform=ax.transAxes,
                ha='center', va='bottom',
                fontsize=11, color=(0/255, 179/255, 176/255, 0.8), style='italic')
        
        ax.grid(axis='x', alpha=0.3, linestyle='--', color='#00B3B0')
        ax.set_axisbelow(True)
        ax.set_xlim(0, max(valores) * 1.18)
        ax.invert_yaxis()
        
        fig.patch.set_facecolor('white')
        ax.set_facecolor('white')
        
        plt.tight_layout()

        caminho_exec, caminho_ultimo = self._caminhos_saida('grafico_evolucao_temporal.png')
        plt.savefig(caminho_exec, dpi=300, bbox_inches='tight', facecolor='white')
        self._atualizar_ultimo(caminho_exec, caminho_ultimo)
        
        self.imagens_base64['grafico_evolucao'] = self._fig_to_base64(fig)
        
        print(f"  ✓ Gráfico de evolução temporal (PNG + Base64)")

    def _grafico_pizza_situacao(self):
        """Gráfico de pizza: distribuição por situação"""
        fig, ax = plt.subplots(figsize=(10, 8))
        
        dados = self.resultados['por_situacao']
        cores = self.config.CORES_PADRAO[:len(dados)]
        
        wedges, texts, autotexts = ax.pie(
            dados.values, 
            labels=dados.index, 
            autopct='%1.1f%%',
            colors=cores,
            startangle=90,
            textprops={'fontsize': 11, 'fontweight': 'bold'}
        )
        
        ax.set_title('Distribuição por Situação', 
                     fontsize=16, fontweight='bold', pad=20)
        
        plt.tight_layout()
        
        caminho_exec, caminho_ultimo = self._caminhos_saida('grafico_distribuicao_situacao.png')
        plt.savefig(caminho_exec, dpi=300, bbox_inches='tight')
        self._atualizar_ultimo(caminho_exec, caminho_ultimo)
        
        self.imagens_base64['grafico_situacao'] = self._fig_to_base64(fig)
        
        print(f"  ✓ Gráfico de distribuição (PNG + Base64)")
    
    def _grafico_top_tipos(self):
        """Gráfico de barras horizontais: top 10 tipos"""
        fig, ax = plt.subplots(figsize=(12, 8))
        
        dados = self.resultados['por_tipo'].sort_values()
        cores = self.config.CORES_PADRAO[:len(dados)]
        
        dados.plot(kind='barh', ax=ax, color=cores, edgecolor='black')
        
        ax.set_title('Top 10 Tipos de Requerimento', 
                     fontsize=16, fontweight='bold', pad=20)
        ax.set_xlabel('Quantidade', fontsize=12, fontweight='bold')
        ax.set_ylabel('Tipo de Requerimento', fontsize=12, fontweight='bold')
        ax.grid(axis='x', alpha=0.3)
        
        for i, v in enumerate(dados.values):
            ax.text(v + 0.5, i, str(v), va='center', fontweight='bold')
        
        plt.tight_layout()
        
        caminho_exec, caminho_ultimo = self._caminhos_saida('grafico_top_tipos.png')
        plt.savefig(caminho_exec, dpi=300, bbox_inches='tight')
        self._atualizar_ultimo(caminho_exec, caminho_ultimo)
        
        self.imagens_base64['grafico_tipos'] = self._fig_to_base64(fig)
        
        print(f"  ✓ Gráfico de top tipos (PNG + Base64)")
    
    def gerar_relatorio_excel(self):
        """Gera relatório consolidado em Excel"""
        print("\n📑 GERANDO RELATÓRIO EXCEL...")

        caminho_exec, caminho_ultimo = self._caminhos_saida('relatorio_consolidado.xlsx')
        
        with pd.ExcelWriter(caminho_exec, engine='openpyxl') as writer:
            df_export = self.dados_processados.copy()
            cols_situacao_manter = {'SITUACAO_ORIGINAL', 'SITUACAO_NORMALIZADA', 'EH_RESPONDIDO', 'COLUNA_SITUACAO_USADA'}
            cols_drop = [c for c in df_export.columns if ('SITUA' in str(c).upper()) and (c not in cols_situacao_manter)]
            if cols_drop:
                df_export = df_export.drop(columns=cols_drop)

            df_export.to_excel(
                writer, 
                sheet_name='Dados Consolidados', 
                index=False
            )
            print("  ✓ Aba 'Dados Consolidados'")
            
            kpis_df = pd.DataFrame({
                'KPI': [
                    'Total de Requerimentos Respondidos',
                    'Colaborador com Maior Volume',
                    'Quantidade do Top Colaborador',
                    'Média Diária'
                ],
                'Valor': [
                    self.resultados.get('total_respondidos', 'N/A'),
                    self.resultados.get('top_colaborador', 'N/A'),
                    self.resultados.get('top_colaborador_qtd', 'N/A'),
                    f"{self.resultados.get('media_dia', 0):.1f}" if 'media_dia' in self.resultados else 'N/A'
                ]
            })
            kpis_df.to_excel(writer, sheet_name='KPIs Resumo', index=False)
            print("  ✓ Aba 'KPIs Resumo'")
            
            # NOVIDADE v3.2: Aba separada por data
            if 'por_colaborador_com_data' in self.resultados:
                collab_data_df = pd.DataFrame({
                    'Colaborador (Data)': self.resultados['por_colaborador_com_data'].index,
                    'Quantidade': self.resultados['por_colaborador_com_data'].values
                })
                collab_data_df.to_excel(writer, sheet_name='Por Colaborador (Data)', index=False)
                print("  ✓ Aba 'Por Colaborador (Data)'")
            
            if self.log_detalhado:
                log_df = pd.DataFrame(self.log_detalhado)
                log_df.to_excel(writer, sheet_name='Log Processamento', index=False)
                print("  ✓ Aba 'Log Processamento'")
        
        self._atualizar_ultimo(caminho_exec, caminho_ultimo)
        print(f"  ✅ Relatório Excel salvo")
    
    def gerar_relatorio_html(self):
        """Gera relatório executivo em HTML"""
        print("\n📄 GERANDO RELATÓRIO HTML (v3.2 - COM SEPARAÇÃO POR DATA)...")
        
        html_content = self._criar_html_relatorio()

        caminho_exec, caminho_ultimo = self._caminhos_saida('relatorio_executivo.html')
        with open(caminho_exec, 'w', encoding='utf-8') as f:
            f.write(html_content)

        try:
            with open(caminho_ultimo, 'w', encoding='utf-8') as f:
                f.write(html_content)
        except Exception as e:
            print(f"  ⚠ Erro ao atualizar ULTIMO")

        print(f"  ✅ Relatório HTML salvo (com imagens embutidas + separação por data)")
        print(f"  📧 Perfeito para enviar por email - arquivo único autocontido!")
    
    def _criar_html_relatorio(self) -> str:
        """Cria o conteúdo HTML do relatório"""
        
        total = self.resultados.get('total_respondidos', 0)
        top_collab = self.resultados.get('top_colaborador', 'N/A')
        top_qtd = self.resultados.get('top_colaborador_qtd', 0)
        media_dia = self.resultados.get('media_dia', 0)
        
        # Tabela de colaboradores COM DATA
        tabela_collab = ""
        if 'por_colaborador_com_data' in self.resultados:
            tabela_collab = "<table style='width:100%; border-collapse: collapse; margin-top: 20px;'>"
            tabela_collab += "<tr style='background-color: #00B3B0; color: white;'>"
            tabela_collab += "<th style='padding: 12px; text-align: left;'>Colaborador (Data)</th>"
            tabela_collab += "<th style='padding: 12px; text-align: center;'>Quantidade</th>"
            tabela_collab += "<th style='padding: 12px; text-align: center;'>Percentual</th>"
            tabela_collab += "</tr>"
            
            for collab, qtd in self.resultados['por_colaborador_com_data'].items():
                perc = (qtd / total * 100) if total > 0 else 0
                tabela_collab += f"<tr style='border-bottom: 1px solid rgba(255,255,255,0.1);'>"
                tabela_collab += f"<td style='padding: 10px; color: white;'>{collab}</td>"
                tabela_collab += f"<td style='padding: 10px; text-align: center; color: white;'><strong>{qtd}</strong></td>"
                tabela_collab += f"<td style='padding: 10px; text-align: center; color: white;'>{perc:.1f}%</td>"
                tabela_collab += "</tr>"
            
            tabela_collab += "</table>"
        
        data_hora = datetime.now().strftime("%d/%m/%Y às %H:%M:%S")
        
        periodo_inicio = 'N/A'
        periodo_fim = 'N/A'
        if 'evolucao_diaria' in self.resultados and len(self.resultados['evolucao_diaria']) > 0:
            periodo_inicio = datetime.strptime(str(self.resultados['evolucao_diaria'].index[0]), '%Y-%m-%d').strftime('%d.%m.%Y')
            periodo_fim = datetime.strptime(str(self.resultados['evolucao_diaria'].index[-1]), '%Y-%m-%d').strftime('%d.%m.%Y')
        
        # Seção de gráficos
        secao_graficos = ""
        
        if 'grafico_colaboradores' in self.imagens_base64:
            secao_graficos += f"""
            <div style="margin: 30px 0; text-align: center;">
                <img src="{self.imagens_base64['grafico_colaboradores']}" 
                     alt="Gráfico de Colaboradores por Data" 
                     style="max-width: 100%; border-radius: 10px; box-shadow: 0 4px 15px rgba(0,0,0,0.3);">
            </div>
            """
        
        if 'grafico_evolucao' in self.imagens_base64:
            secao_graficos += f"""
            <div style="margin: 30px 0; text-align: center;">
                <img src="{self.imagens_base64['grafico_evolucao']}" 
                     alt="Gráfico de Evolução Temporal" 
                     style="max-width: 100%; border-radius: 10px; box-shadow: 0 4px 15px rgba(0,0,0,0.3);">
            </div>
            """
        
        if 'grafico_situacao' in self.imagens_base64:
            secao_graficos += f"""
            <div style="margin: 30px 0; text-align: center;">
                <img src="{self.imagens_base64['grafico_situacao']}" 
                     alt="Gráfico de Distribuição por Situação" 
                     style="max-width: 100%; border-radius: 10px; box-shadow: 0 4px 15px rgba(0,0,0,0.3);">
            </div>
            """
        
        if 'grafico_tipos' in self.imagens_base64:
            secao_graficos += f"""
            <div style="margin: 30px 0; text-align: center;">
                <img src="{self.imagens_base64['grafico_tipos']}" 
                     alt="Gráfico Top 10 Tipos" 
                     style="max-width: 100%; border-radius: 10px; box-shadow: 0 4px 15px rgba(0,0,0,0.3);">
            </div>
            """
        
        html = f"""
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Relatório Executivo v3.2 - Gestão de Requerimentos</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: linear-gradient(135deg, #061A2B 0%, #0A2A44 100%);
            color: white;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: rgba(255,255,255,0.05);
            padding: 40px;
            border-radius: 15px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.3);
        }}
        h1 {{
            color: #00B3B0;
            text-align: center;
            margin-bottom: 10px;
            font-size: 2.2em;
        }}
        h2 {{
            color: #3EE7DA;
            border-bottom: 2px solid #00B3B0;
            padding-bottom: 10px;
            margin-top: 30px;
        }}
        .subtitle {{
            text-align: center;
            color: #82C0CC;
            margin-bottom: 30px;
            font-size: 1em;
        }}
        .kpi-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 20px;
            margin: 30px 0;
        }}
        .kpi-box {{
            background: linear-gradient(135deg, #00B3B0 0%, #16697A 100%);
            padding: 25px;
            border-radius: 12px;
            text-align: center;
            box-shadow: 0 4px 15px rgba(0,179,176,0.3);
            transition: transform 0.3s;
        }}
        .kpi-box:hover {{
            transform: translateY(-5px);
        }}
        .kpi-label {{
            font-size: 0.9em;
            margin-bottom: 10px;
            opacity: 0.9;
        }}
        .kpi-value {{
            font-size: 2.2em;
            font-weight: bold;
            color: white;
        }}
        .badge {{
            display: inline-block;
            padding: 5px 15px;
            background: rgba(62,231,218,0.2);
            border-radius: 20px;
            font-size: 0.85em;
            margin: 5px;
        }}
        .footer {{
            text-align: center;
            margin-top: 50px;
            padding-top: 20px;
            border-top: 1px solid rgba(255,255,255,0.1);
            color: #82C0CC;
            font-size: 0.9em;
        }}
        table {{
            font-size: 0.95em;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>⚡ Relatório Executivo v3.2</h1>
        <div class="subtitle">
            <strong>Sistema de Gestão de Requerimentos - CSC</strong><br>
            Gerado em: {data_hora}<br>
            <span class="badge">📅 Período: {periodo_inicio} a {periodo_fim}</span>
            <span class="badge">🔑 Separado por Data</span>
            <span class="badge">🖼️ Imagens embutidas</span>
        </div>
        
        <h2>📊 KPIs Principais</h2>
        <div class="kpi-grid">
            <div class="kpi-box">
                <div class="kpi-label">Total Respondidos</div>
                <div class="kpi-value">{total}</div>
            </div>
            <div class="kpi-box">
                <div class="kpi-label">Top Colaborador+Data</div>
                <div class="kpi-value" style="font-size: 1.3em;">{top_collab}</div>
                <div style="margin-top: 5px; font-size: 0.9em;">({top_qtd} requerimentos)</div>
            </div>
            <div class="kpi-box">
                <div class="kpi-label">Média Diária</div>
                <div class="kpi-value">{media_dia:.1f}</div>
                <div style="margin-top: 5px; font-size: 0.9em;">req/dia</div>
            </div>
        </div>
        
        <h2>👥 Produtividade por Colaborador (Separado por Data)</h2>
        <p style="color: #82C0CC; font-size: 0.9em; margin-top: 10px;">
            ℹ️ Cada linha representa um colaborador em uma data específica (não soma datas diferentes)
        </p>
        {tabela_collab}
        
        <h2>📈 Visualizações</h2>
        {secao_graficos}
        
        <div class="footer">
            <p><strong>🚀 Sistema de Análise v3.2</strong></p>
            <p>✨ Colaboradores separados por data - Imagens embutidas - Perfeito para email</p>
            <p style="margin-top: 10px; font-size: 0.85em; opacity: 0.7;">
                Exemplo: "Andrey (05.02.2026)" e "Andrey (06.02.2026)" são contabilizados separadamente
            </p>
        </div>
    </div>
</body>
</html>
"""
        return html
    
    def executar_analise_completa(self):
        """Executa o pipeline completo de análise"""
        try:
            self.carregar_planilhas()
            self.processar_dados()
            self.calcular_kpis()
            self.gerar_graficos()
            self.gerar_relatorio_excel()
            self.gerar_relatorio_html()
            
            print("\n" + "=" * 80)
            print("✅ ANÁLISE CONCLUÍDA COM SUCESSO! (v3.2 - SEPARAÇÃO POR DATA)")
            print("=" * 80)
            print(f"\n📁 Resultados em: {self.pasta_saida_execucao}")
            print("\n📄 Arquivos gerados:")
            print("  • relatorio_consolidado.xlsx")
            print("  • relatorio_executivo.html (⭐ IMAGENS EMBUTIDAS + SEPARAÇÃO POR DATA)")
            print("  • grafico_colaboradores_por_data.png")
            print("  • grafico_evolucao_temporal.png")
            print("  • grafico_distribuicao_situacao.png")
            print("  • grafico_top_tipos.png")
            print("\n🔑 IMPORTANTE: Colaboradores de datas diferentes NÃO são somados")
            print("   Exemplo: Andrey (05.02) ≠ Andrey (06.02)")
            print("\n📧 O arquivo HTML está pronto para envio por email!")
            print("=" * 80)
            
        except Exception as e:
            print(f"\n❌ ERRO: {type(e).__name__}: {str(e)}")
            import traceback
            traceback.print_exc()


# ============================================================================
# EXECUÇÃO PRINCIPAL
# ============================================================================

if __name__ == "__main__":
    print("\n🚀 Iniciando Sistema de Análise v3.2 (SEPARAÇÃO POR DATA)\n")
    
    config = Config()
    analisador = AnalisadorRequerimentos(config)
    analisador.executar_analise_completa()
    
    print("\n✨ Processo finalizado!\n")